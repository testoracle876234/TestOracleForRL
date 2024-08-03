import numpy as np
import control
from scipy import linalg as la
import openpyxl

def trueCount(array, metric):
    counter = 0
    for i in range(20):
        count = np.sum(array[i][metric])
        counter += count
    return counter

def trueCountAll(array, metric, beta, num_cases):
    counter = 0
    for i in range(20):
        count = np.sum(array[i][metric])
        if count/num_cases >= beta:
            counter += 1
    return counter

def save_to_excel(filename, sheetname, data_list, column):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheetname]

    start_row = 2

    if sheetname == 'LPEA':
        for i in range(0, 6):
            cell = sheet.cell(row=start_row + i, column=column)
            cell.value = data_list[i]    
    elif sheetname == 'DSC':
        for i in range(8):
            cell = sheet.cell(row=start_row + i, column=column)
            cell.value = data_list[i]    
    wb.save(filename)

def get_data_list(metrics, num_cases):
    data_list =  [
        trueCount(metrics, 1),
        trueCountAll(metrics, 1, 1, num_cases),
        trueCountAll(metrics, 1, 0.975, num_cases),
        trueCountAll(metrics, 1, 0.95, num_cases),
        trueCountAll(metrics, 1, 0.925, num_cases),
        trueCountAll(metrics, 1, 0.9, num_cases),
    ]
    return data_list

def findPK(A, B, Q, R):
        K = control.dlqr(A, B, Q, R)[0]
        Acl = A - B @ K
        P = la.solve_discrete_lyapunov(Acl, Q)
        is_positive_definite = np.all(np.linalg.eigvals(P) > 0)
        if is_positive_definite:
            print("System can be stable")
            return K, P
        else:
            print("System can't be stable")
            return None, None

def generate_trajectory_for_model(env, model, X0, num_steps):
    X = np.zeros((num_steps, len(X0), 1)) 
    X[0] = X0.reshape((2, 1))
    U = np.zeros((num_steps, len(X0), 1))

    for i in range(1, num_steps):
        X[i-1] = X[i-1].reshape((2, 1))
        action = model.predict(X[i-1].reshape((2, 1)), deterministic=True)[0]
        X[i] = (env.A @ X[i-1] + env.B @ action).astype(np.float32)
        U[i- 1] = X[i] - X[i-1] 

    return X[:, 0], X[:, 1], U[:, 0], U[:, 1]

def generate_trajectory(A, X0, num_steps):
    X = np.zeros((num_steps, len(X0))) 
    X[0] = X0
    U = np.zeros((num_steps, len(X0)))

    for i in range(1, num_steps):
        X[i] = A @ X[i-1] 
        U[i- 1] = X[i] - X[i-1] 

    return X[:, 0], X[:, 1], U[:, 0], U[:, 1]

def generate_X0_circle(num_points, center, radius):
    X0 = np.zeros((num_points, 2))

    for i in range(num_points):
        angle = i * (2 * np.pi / num_points)
        x = center[0] + radius * np.cos(angle)
        y = center[1] + radius * np.sin(angle)

        X0[i] = [x, y]

    return X0

def generate_field_grid(A, points):
    X = np.zeros((len(points), 2, 1)) 
    U = np.zeros((len(points), 2, 1))
    
    for i, point in enumerate(points):
        X[i] = point
        U[i] = A @ point  - point 
    return X[:, 0], X[:, 1], U[:, 0], U[:, 1]

def generate_field_grid_for_model(model, points, env):
    X = np.zeros((len(points), 2, 1)) 
    U = np.zeros((len(points), 2, 1))
    
    for i, point in enumerate(points):
        X[i] = point
        action = model.predict(X[i].reshape((2, 1)), deterministic=True)[0]
        U[i] = (env.A @ X[i] + env.B @ action).astype(np.float32)  - point 
    return X[:, 0], X[:, 1], U[:, 0], U[:, 1]

def generate_X0_random(num_points, range):
    X0 = np.random.random((num_points, 2)) * range
    return X0

def generate_X0_uniform(num_points, side_length):
    x = np.linspace(-side_length/2, side_length/2, num_points)
    y = np.linspace(-side_length/2, side_length/2, num_points)
    xv, yv = np.meshgrid(x, y)
    coordinates = np.dstack((xv, yv)).reshape(-1, 2, 1)
    return coordinates

def evaluate(model, total_time, env, n):
    env.reset()
    cumulated_punish = 0
    for i in range(total_time):
        env.state = np.random.uniform(-1, 1, size=(n, 1))
        state = env.state
        ideal_action = - env.K @ state
        action = model.predict(state, deterministic=True)
        cumulated_punish += abs(action[0] - ideal_action).sum()
    return cumulated_punish/total_time

def generate_state_transition_matix(n, m):
    is_A_stable = True
    while(is_A_stable):
        A = np.random.rand(n, n)
        eigenvalues = np.linalg.eigvals(A)
        is_A_stable = all(np.abs(eigenvalues) < 1)
    B = np.random.rand(n, m)
    return A, B

def generate_points_around(center, d):
    center = center.reshape((2,))
    num_dimensions = len(center)
    num_points = 2 * num_dimensions

    points = np.zeros((num_points, num_dimensions))

    for i in range(num_dimensions):
        points[i] = center.copy()
        points[i][i] += d

        points[num_dimensions + i] = center.copy()
        points[num_dimensions + i][i] -= d

    return points

def generate_action_field_grid_for_model(model, points):
    X = np.zeros((len(points), 2, 1)) 
    U = np.zeros((len(points), 2, 1))
    
    for i, point in enumerate(points):
        X[i] = point
        action = model.predict(X[i].reshape((2, 1)), deterministic=True)[0]
        U[i] = action
    return X[:, 0], X[:, 1], U[:, 0], U[:, 1]

def generate_action_field_grid(env, points):
    X = np.zeros((len(points), 2, 1)) 
    U = np.zeros((len(points), 2, 1))
    
    for i, point in enumerate(points):
        X[i] = point
        U[i] = -env.K @ point
    return X[:, 0], X[:, 1], U[:, 0], U[:, 1]
